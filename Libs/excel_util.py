#coding=utf-8
from openpyxl import load_workbook
import sys

class ReadExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self):
        wb =load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        titles = []
        for i in range(1,sheet.max_column + 1):
            titles.append(sheet.cell(1,i).value)

        all_data = []
        for i in range(2,sheet.max_row + 1):
            row_data = {}
            for j in range(1,sheet.max_column +1):
                row_data[titles[j-1]] = sheet.cell(i,j).value
            all_data.append(row_data)

        return all_data

    def excel_case_switch(self,tcname):
        cases = self.read_excel()
        for i in range(len(cases)):
            if cases[i]['TestCaseName'] == tcname:
                if cases[i]['Flag'].lower() == 'on':
                    return True
                else:
                    return False
        else:
            return False

# d = ReadExcel(sys.path[1] + '\\Data\\TestCaseON-OFF.xlsx','smokeTest')
# print(d.read_excel())